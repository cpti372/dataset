#엑셀 데이터 불러오기- 리스트로도 읽고, 그냥으로도 출력
from openpyxl import load_workbook
read_workbook=load_workbook("./bluehouse.xlsx")
read_cell=read_workbook.active
result_list=[]
for i in range(1,15):
    print(read_cell.cell(i,1).value)
    result_list.append(read_cell.cell(i,1).value)
print(result_list)