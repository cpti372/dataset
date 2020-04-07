#청와대 국민청원 홈페이지 크롤링하기 
from openpyxl import load_workbook

read_wb=load_workbook("./bluehouse.xlsx")
read_ws=read_wb.active

list_excel=[]
for i in range(1,15):
    print(read_ws.cell(i,1).value)
    list_excel.append(read_ws.cell(i,1).value)

print(list_excel)

from konlpy.tag import Kkma
from konlpy.utils import pprint
kkma = Kkma()
import collections
list_temp=[]
for row in list_excel:
    print(kkma.nouns(row))
    list_temp=list_temp+kkma.nouns(row)
print(list_temp)

result_list=[]
for check in list_temp:
    print(check)
    if len(check)>1:
        result_list.append(check)
print(collections.Counter(result_list))
print("------------------------------------")
print(collections.Counter(result_list).most_common(10))