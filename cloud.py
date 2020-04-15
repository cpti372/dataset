#청와대 국민청원 홈페이지 크롤링하기 
from openpyxl import load_workbook
from wordcloud import WordCloud
read_wb=load_workbook("./bluehouse.xlsx")
read_ws=read_wb.active

list_excel=[]
for i in range(1,15):
    print(read_ws.cell(i,1).value)
    list_excel.append(read_ws.cell(i,1).value)

print(list_excel)

from konlpy.tag import Kkma
from konlpy.utils import pprint
kkma = Kkma()
import collections
list_temp=[]
for row in list_excel:
    print(kkma.nouns(row))
    list_temp=list_temp+kkma.nouns(row)
print(list_temp)

result_list=[]
for check in list_temp:
    print(check)
    if len(check)>1:
        result_list.append(check)

last_text=" "
for i in result_list:
    last_text= last_text+" "+i
print(last_text)
import matplotlib.pyplot as plt
wordCloud= WordCloud(font_path="/Users/kimyeonsu/Downloads/NotoSansCJKkr-Medium.otf").generate(last_text)
wordCloud.words_
plt.imshow(wordCloud,interpolation="bilinear")
plt.axis("off")
plt.show()

